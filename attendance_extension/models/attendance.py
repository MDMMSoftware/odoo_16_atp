from odoo import fields, models, api, _
from datetime import timedelta,datetime
from odoo.exceptions import ValidationError,UserError
import pytz
from geopy.geocoders import Nominatim
import logging
import base64
import pandas as pd
from io import BytesIO
from collections import namedtuple
from odoo.addons.resource.models.resource import float_to_time
from openpyxl import load_workbook

from pytz import timezone, UTC
_logger = logging.getLogger(__name__)
DummyAttendance = namedtuple('DummyAttendance', 'hour_from, hour_to, dayofweek, day_period, week_type')

class HrAttendance(models.Model):
    _inherit = "hr.attendance"

    latti_in = fields.Char(string='Latitude In')
    longi_in = fields.Char(string='Longitube In')
    latti_out = fields.Char(string='Latitude Out')
    longi_out = fields.Char(string='Longitube Out')
    check_in_address = fields.Char(string='Check In Address')
    check_out_address = fields.Char(string='Check Out Address')
    department_id = fields.Many2one('hr.department', string="Department", related="employee_id.department_id",
        readonly=True,store=True)

    employee_id_code = fields.Char('Employee Code',related="employee_id.employee_id")
    branch_id = fields.Many2one('hr.branch', related="employee_id.branch_id")
    job_location_id = fields.Many2one('hr.job.location')
    # check_in_job_location_id = fields.Many2one('hr.job.location')
    # check_out_job_location_id = fields.Many2one('hr.job.location')
    
    check_in_address = fields.Char(string='Address')
    check_out_address = fields.Char(string='Address')
    checkin_location = fields.Char(string='Link', store=True,help="Check in location link of the User")
    checkout_location = fields.Char(string='Link', store=True,help="Check out location link of the User")
    auto_checkout = fields.Boolean(default=False,string="Auto Checkout")

    check_in_date = fields.Date(string="Checkin Date",compute="compute_check_in_out_date",store=True)
    check_out_date =  fields.Date(string="Checkout Date",compute="compute_check_in_out_date",store=True)

    @api.onchange('employee_id')
    def _onchange_employee_id(self):
        for rec in self:
            if rec.employee_id:
                rec.job_location_id = rec.employee_id.job_location

    @api.depends('check_in_date', 'check_out_date')
    def compute_check_in_out_date(self):
        for rec in self:
            if rec.check_in:
                rec.check_in_date = rec.check_in.date()
            if rec.check_out:
                rec.check_out_date = rec.check_out.date()

    def unlink(self):
        for _ in self:
            raise UserError("Are you doing something fraudly! Why do you want to delete some records?? 🤔 ")
        return super().unlink() 

    def get_own_attendance(self, emp):
        """
            --> API <--
            Get Own attendances
        """
        current_date = fields.Datetime.now()

        # Get start date of current month with 0 hours and 0 minutes
        start_of_current_month = current_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        # Get end date of current month with 23 hours and 59 minutes
        end_of_current_month = start_of_current_month.replace(
            month=start_of_current_month.month % 12 + 1,
            year=start_of_current_month.year + (1 if start_of_current_month.month == 12 else 0),
            day=1,
            hour=0,
            minute=0,
            second=0,
            microsecond=0
        ) - timedelta(minutes=1)
        all_attendances = self.env['hr.attendance'].search([('employee_id','=',int(emp)),('check_in','>=',start_of_current_month.strftime("%Y-%m-%d %H:%M")),('check_in','<=',end_of_current_month.strftime("%Y-%m-%d %H:%M"))])
        lst = []
        for attendance in all_attendances:
            temp = [False,False,False]
            if attendance.check_in:
                temp[0] = datetime.date(attendance.check_in)
                temp[1] = datetime.time(attendance.check_in + timedelta(hours=6, minutes=30))
            if attendance.check_out:
                temp[2] = datetime.time(attendance.check_out + timedelta(hours=6, minutes=30))
            lst.append(temp)
        return lst
    
    def _is_working_day(self, calendar, date):
        weekday = date.weekday()
        is_working = any(attendance.dayofweek == str(weekday) for attendance in calendar.attendance_ids)
        is_public_holiday = self.env['resource.calendar.leaves'].search_count([
            ('date_from', '<=', date),
            ('date_to', '>=', date),
            ('resource_id','=',False)
        ]) > 0
        
        return is_working and not is_public_holiday


    def attendance_action(self, latt, long, uid, emp, checker = False):
        """ 
            --> API <--
            Check In/Check Out action
            Check In: create a new attendance record
            Check Out: modify check_out field of appropriate attendance record
        """
        # action date to write attendance date - with odoo fields.datetime
        action_date = fields.Datetime.now()    
        # datetimes to check attendance action is valid or not - with python datetime module
        check_action_date = action_date + timedelta(hours=6,minutes=30,seconds=00)  
        check_start_date =check_action_date.replace(hour=0,minute=0,second=0)
        check_end_date = check_start_date+timedelta(hours=23,minutes=59,seconds=59)    
        # logged the action
        _logger.info(f"Employee ID {emp} is trying to make attendance action at UTC+6:30 {check_action_date} , server-tz {action_date}")
        
        # check the current employee has check in or check out with the check_action_date
        checkin_attendance = self.env['hr.attendance'].search([('employee_id', '=', int(emp)), ('check_in', '>', check_start_date),('check_in', '<', check_end_date)])
        checkout_attendance = self.env['hr.attendance'].search([('employee_id', '=', int(emp)), ('check_out', '>', check_start_date),('check_out', '<', check_end_date)])
        
        no_checkout_attendance = self.env['hr.attendance'].search([('employee_id','=', int(emp)), ('check_out', '=',False)],limit=1)
        
        # return the attendance values if the current action is just checker
        if checker:
            if no_checkout_attendance:
                return [int(emp), no_checkout_attendance.id, False]
            return [int(emp), checkin_attendance.id, checkout_attendance.id]
        
        # check the holiday check in with check_action_date
        check_in_allow = self._is_working_day(self.env['hr.employee'].browse(emp).resource_calendar_id,check_action_date)
        if not check_in_allow:
            raise ValidationError(_("You don't have to attend on holiday day."))        
        
        # # if it is not just checker
        # if both check in and check out existed , don't allow
        if checkin_attendance and checkout_attendance:
            raise ValidationError(_("You have already checked in once or checked out once"))
        # if only checked in , but no checkout 
        elif (checkin_attendance and not checkout_attendance) or no_checkout_attendance:
            attendance = self.env['hr.attendance'].search([('employee_id', '=', int(emp)), ('check_out', '=', False)], limit=1)            
            if attendance:
                # get check in time , reject check out if it is within one hour of check in
                check_in_time = attendance.check_in
                if (action_date - check_in_time) < timedelta(hours=1,minutes=0,seconds=0):
                    raise ValidationError(_("You can't check out within 1 hour of check in time"))
                # write the action server date as check out
                attendance.sudo().write({
                    'latti_out': latt,
                    'longi_out': long,
                    'write_uid':uid,
                    'write_date':action_date,
                    'check_out': action_date,
                })
                _logger.info(f"Employee ID {emp} is successfully checked out at UTC+6:30 {check_action_date} , server-tz {action_date}")
                return 'check_in'
            raise ValidationError(_("You haven't checked in yet!!"))
        # if no check in and no checkout
        elif not checkin_attendance and not checkout_attendance:
            # if the actual check in action date is over 2 PM , reject the action
            if check_action_date.hour>=14:
                raise ValidationError(_("You can't check in at/over 2 PM."))   
            emp_obj = self.env['hr.employee'].browse(emp)
            # create the attendance with the action server date as check in
            vals = {
                'employee_id':emp,
                'job_location_id':emp_obj.job_location.id,
                'latti_in': latt,
                'longi_in': long,
                'create_uid':uid,
                'check_in':action_date,
                'create_date':check_action_date
            }
            attendance = self.env['hr.attendance'].sudo().create(vals)
            if attendance:
                self.env.cr.execute(
                        'UPDATE hr_attendance SET create_date=%s WHERE id=%s',
                        (action_date+ timedelta(hours=6,minutes=30,seconds=00),attendance.id,)
                    )
            _logger.info(f"Employee ID {emp} is successfully checked in at UTC+6:30 {check_action_date} , server-tz {action_date}")
            return 'check_out'
        else:
            raise ValidationError('Unknown Error')
        
    # Scheduled Actions
        
    # 1. Auto Checkout
    def _auto_checkout(self):
        no_checkout_line = self.env['hr.attendance'].search([('check_out','=',False)])
        for line in no_checkout_line:
            if line.check_in.strftime("%A")=='Saturday':
                if line.check_in < line.check_in.replace(hour=16,minute=0,second=0):
                    line.write({'check_out':line.check_in.replace(hour=16,minute=0,second=0)-timedelta(hours=6,minutes=30),
                                'latti_out':line.latti_in,
                                'longi_out':line.longi_in,
                                'check_out_address':line.check_in_address,
                                'checkout_location':line.checkin_location,
                                'auto_checkout':True})
            else:
                if line.check_in < line.check_in.replace(hour=16,minute=30,second=0):
                    line.write({'check_out':line.check_in.replace(hour=16,minute=30,second=0)-timedelta(hours=6,minutes=30),
                                'latti_out':line.latti_in,
                                'longi_out':line.longi_in,
                                'check_out_address':line.check_in_address,
                                'checkout_location':line.checkin_location,
                                'auto_checkout':True})   

    # 2. Auto Generate Address from Coordinates
    def _auto_generate_map(self):
        no_location_line = self.env['hr.attendance'].search(['|',('check_in_address','=',False),('check_out_address','=',False)],limit=500)
        for line in no_location_line:
            if line.latti_in and line.longi_in:
                lt = line.latti_in
                lg = line.longi_in
                geolocator = Nominatim(user_agent='my-app')
                location = geolocator.reverse(lt+","+lg)
                line.write({'check_in_address': location.address,
                            'checkin_location': f"https://www.google.com/maps/search/?api=1&query={lt},{lg}"
                            })
            if line.latti_out and line.longi_out:
                lt_out = line.latti_out
                lg_out = line.longi_out
                geolocator = Nominatim(user_agent='my-app')
                location = geolocator.reverse(lt_out+","+lg_out)
                line.write({'check_out_address': location.address,
                            'checkout_location': f"https://www.google.com/maps/search/?api=1&query={lt},{lg}"}) 

class HrEmployee(models.AbstractModel):
    """Inherits HR Employee model"""
    _inherit = 'hr.employee'                  

    ## Overriding Attendances Actions 
    def attendance_manual(self, next_action, entered_pin=None):
        """Override this method to add latitude and longitude"""
        self.ensure_one()
        
        # logged the action     
        _logger.info(f"Employee ID {self.id} is trying to make attendance action at UTC+6:30 {fields.Datetime.now() + timedelta(hours=6,minutes=30,seconds=00)  } , server-tz {fields.Datetime.now()}")

        latitudes = self.env.context.get('latitude', False)
        longitudes = self.env.context.get('longitude', False)
        print("latt",latitudes,"----",longitudes)
        if not latitudes or not longitudes:
            raise ValidationError("Lattitude and longitude are not found!!!")
        attendance_user_and_no_pin = self.user_has_groups('hr_attendance.group_hr_attendance_user,'
                                                          '!hr_attendance.group_hr_attendance_use_pin')
        can_check_without_pin = attendance_user_and_no_pin or self.sudo().user_id == self.env.user and entered_pin is None

        if can_check_without_pin or entered_pin is not None and entered_pin == self.sudo().pin:
            return self._attendance_action(latitudes, longitudes,next_action)
        if not self.user_has_groups('hr_attendance.group_hr_attendance_user'):
            return {'warning': _('To activate Kiosk mode without pin code, you '
                                 'must have access right as an Officer or above'
                                 'in the Attendance app. Please contact your '
                                 'administrator.')}
        return {'warning': _('Wrong PIN')}

    def _is_working_day(self, calendar, date):
        weekday = date.weekday()
        is_working = any(attendance.dayofweek == str(weekday) for attendance in calendar.attendance_ids)
        is_public_holiday = self.env['resource.calendar.leaves'].search_count([
            ('date_from', '<=', date),
            ('date_to', '>=', date),
            ('resource_id','=',False)
        ]) > 0
        
        return is_working and not is_public_holiday

    def _attendance_action(self, latitudes, longitudes,next_action):
        """ Changes the attendance of the employee.
            Returns an action to the check in/out message,
            next_action defines which menu the check in/out message should
            return to. ("My Attendances" or "Kiosk Mode")
        """
        check_in_allow = self._is_working_day(self.resource_calendar_id,fields.Datetime.now())
        if not check_in_allow:
            raise ValidationError(_("You don't have to attend on holiday day."))

        self.ensure_one()
        employee = self.sudo()
        action_message = self.env['ir.actions.actions']._for_xml_id('hr_attendance.'
                                                                    'hr_attendance_action_greeting_message')

        action_message['previous_attendance_change_date'] = employee.last_attendance_id and (
                employee.last_attendance_id.check_out
                or employee.last_attendance_id.check_in) or False
        action_message['employee_name'] = employee.name
        action_message['barcode'] = employee.barcode
        action_message['next_action'] = next_action
        action_message['hours_today'] = employee.hours_today
        
        if employee.user_id:
            modified_attendance = employee.with_user(employee.user_id).sudo()._attendance_action_change(longitudes,latitudes)
        else:
            modified_attendance = employee._attendance_action_change(longitudes,latitudes)
        action_message['attendance'] = modified_attendance.read()[0]

        return {'action': action_message}

    def _attendance_action_change(self, longitudes, latitudes):
        """ Check In/Check Out action
            Check In: create a new attendance record
            Check Out: modify check_out field of appropriate attendance record
        """
        self.ensure_one()
        action_date = fields.Datetime.now()
        start_date =action_date.replace(hour=0,minute=0,second=0)
        end_date = start_date+timedelta(hours=23,minutes=59,seconds=59)
        checkin_attendance = self.env['hr.attendance'].search([('employee_id', '=', self.id), ('check_in', '>', start_date),('check_in', '<', end_date)])
        checkout_attendance = self.env['hr.attendance'].search([('employee_id', '=', self.id), ('check_out', '>', start_date),('check_out', '<', end_date)])
        if (self.attendance_state != 'checked_in' and checkin_attendance) or checkout_attendance:
            raise ValidationError(_("You have alredy checked in / checked out!"))
        job_location = self.job_location.id
        if self.attendance_state != 'checked_in':
            vals = {
                'employee_id': self.id,
                'job_location_id': job_location,
                'latti_in': latitudes,
                'longi_in': longitudes,
            }
            now = fields.Datetime.now()+timedelta(hours=6,minutes=30)
            if now.hour>=14:
                raise ValidationError(_("You can't check in at/over 2 PM."))
            _logger.info(f"Employee ID {self.id} is successfully checked in {fields.Datetime.now() + timedelta(hours=6,minutes=30,seconds=00)} , server-tz {fields.Datetime.now()}")
            return self.env['hr.attendance'].create(vals)
        attendance = self.env['hr.attendance'].search([('employee_id', '=', self.id), ('check_out', '=', False)], limit=1)
        if attendance:
            check_in_time = attendance.check_in
            if (action_date - check_in_time) < timedelta(hours=1,minutes=0,seconds=0):
                raise ValidationError(_("You can't check out within 1 hour of check in time"))

            attendance.write({
                'latti_out': latitudes,
                'longi_out': longitudes,
            })
            attendance.check_out = action_date
            _logger.info(f"Employee ID {self.id} is successfully checked out {fields.Datetime.now() + timedelta(hours=6,minutes=30,seconds=00)} , server-tz {fields.Datetime.now()}")
        else:
            raise UserError(_('Cannot perform check out on '
                                         '%(empl_name)s, could not find corresponding check in.'
                                         ' Your attendances have probably been modified manually by'
                                         ' human resources.') % {
                                           'empl_name': self.sudo().name})
        return attendance                      
class ChangePasswordUser(models.TransientModel):
    """ A model to configure users in the change password wizard. """
    _inherit = 'change.password.user'
    _description = 'User, Change Password Wizard'

    def change_password_button_hr_app(self,new_passwd,userId):
        self.env['res.users'].search([('id','=',int(userId))]).sudo()._change_password(new_passwd)       
        
class AttendanceImport(models.Model):
    _name = "attendance.import"
    _description = "Attendance Import"
    
    name = fields.Char(string="Description")
    import_date = fields.Date(string='Import Date', readonly=True, default=fields.Date.today())
    import_fname = fields.Char(string='Filename')
    import_file = fields.Binary(string='File', required=True)
    note = fields.Text(string='Log')
    state = fields.Selection([('draft', 'Draft'),('completed', 'Completed'),('error', 'Error')], string='States', default='draft')

    def _check_file_ext(self):
        for import_file in self.browse(self.ids):
            if '.xls' or '.xlsx' in import_file.import_fname:return True
            else: return False
        return True
    
    _constraints = [(_check_file_ext, "Please import EXCEL file!", ['import_fname'])]

    def import_data(self):
        
        import_file = self.import_file
        if not import_file:
            return
        
        file_data = base64.b64decode(import_file)
        file_io = BytesIO(file_data)

        workbook = load_workbook(file_io)
        sheet = workbook.active 
        rows_list = []
        staff_ids = []
        # for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=2, max_col=2):
        #     staff_ids.append(row[0].value)

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            row_values = []
            for cell in row:
                if type(cell.value) == type('str') and cell.value.startswith('MD-'):
                    staff_ids.append(cell.value)
                row_values.append(cell.value)
                
            rows_list.append(row_values)
        date_tuple = rows_list[0]
        time_tuple = rows_list[1]
        date_time_list = list(zip(date_tuple , time_tuple ))
        rows_list = rows_list[2:]
        
        i = 4
        created_count = 0
        for value in date_time_list[4:]:
            
            date = value[0]
            time = value[1]
            for dt_row in rows_list:
                hr_attendance = self.env['hr.attendance']

                employee_id = self.env['hr.employee'].search([('employee_id','=',dt_row[1])])
                if not employee_id:
                    raise UserError(_('Employee %s is not exist.'% dt_row[1]))
                get_attendences = self._get_attendances(employee_id, date.date(), date.date())
                employee_working_hour_dict = {
                    'morning': [get_attendences[0].hour_from,get_attendences[0].hour_to],
                    'afternoon': [get_attendences[1].hour_from,get_attendences[1].hour_to]
                }   
                att_time = self._get_start_or_end_from_attendance(employee_working_hour_dict['morning'][0], date, employee_id) if time == 'Morning' else self._get_start_or_end_from_attendance(employee_working_hour_dict['afternoon'][1], date, employee_id)
                att = {
                    'employee_id': employee_id.id
                }
                if time == 'Morning':
                    if dt_row[i]:
                        job_location = self.env['hr.job.location'].search([('name','=',dt_row[i])],limit=1)
                        if not job_location:
                            raise UserError(_('Location %s is not exist.'% dt_row[i]))

                        att['job_location_id'] = job_location.id
                        att['check_in'] = att_time
                        
                        hr_attendance.create(att)
                        created_count += 1
                elif time == 'Afternoon':
                    if dt_row[i]:
                        job_location = self.env['hr.job.location'].search([('name','=',dt_row[i])],limit=1)
                        if not job_location:
                            raise UserError(_('Location %s is not exist.'% dt_row[i]))

                        check_check_in = self.env['hr.attendance'].search([
                            ('employee_id','=',employee_id.id),
                                ('check_in','=', self._get_start_or_end_from_attendance(employee_working_hour_dict['morning'][0], date, employee_id))
                             ])
                        if not check_check_in:
                            att['job_location_id'] = job_location.id
                            att['check_in'] = self._get_start_or_end_from_attendance(employee_working_hour_dict['afternoon'][0], date, employee_id)
                            att['check_out'] = att_time
                            hr_attendance.create(att)
                            created_count += 1
                            # return UserError(_('There is attendance in %s of %s' % (datetime.strftime(date, "%Y-%m-%d"),employee_id.name)))
                        else:
                            check_check_in.check_out = att_time      
                    else:
                        check_check_in = self.env['hr.attendance'].search([
                            ('employee_id','=',employee_id.id),
                                ('check_in','=', self._get_start_or_end_from_attendance(employee_working_hour_dict['morning'][0],date, employee_id))
                            ])
                        if check_check_in:
                            check_check_in.check_out = self._get_start_or_end_from_attendance(employee_working_hour_dict['morning'][1],date, employee_id)
                        
                
                
            i += 1
        message = 'Import Success at ' + str(datetime.strptime(datetime.today().strftime('%Y-%m-%d %H:%M:%S'),
                '%Y-%m-%d %H:%M:%S'))+ '\n' + str(created_count)+' records imported' +'\
                \n' + str(created_count) + ' Created'
        self.write({'state': 'completed','note': message})

            
    def _get_start_or_end_from_attendance(self, hour, date, employee):
        hour = float_to_time(float(hour))
        holiday_tz = timezone(employee.tz or self.env.user.tz or 'UTC')
        return holiday_tz.localize(datetime.combine(date, hour)).astimezone(UTC).replace(tzinfo=None)

    def _get_attendances(self, employee, request_date_from, request_date_to):
        resource_calendar_id = employee.resource_calendar_id or self.env.company.resource_calendar_id
        domain = [('calendar_id', '=', resource_calendar_id.id), ('display_type', '=', False)]
        attendances = self.env['resource.calendar.attendance'].read_group(domain,
            ['ids:array_agg(id)', 'hour_from:min(hour_from)', 'hour_to:max(hour_to)',
             'week_type', 'dayofweek', 'day_period'],
            ['week_type', 'dayofweek', 'day_period'], lazy=False)

        # Must be sorted by dayofweek ASC and day_period DESC
        attendances = sorted([DummyAttendance(group['hour_from'], group['hour_to'], group['dayofweek'], group['day_period'], group['week_type']) for group in attendances], key=lambda att: (att.dayofweek, att.day_period != 'morning'))

        default_value = DummyAttendance(0, 0, 0, 'morning', False)

        if resource_calendar_id.two_weeks_calendar:
            start_week_type = self.env['resource.calendar.attendance'].get_week_type(request_date_from)
            attendance_actual_week = [att for att in attendances if att.week_type is False or int(att.week_type) == start_week_type]
            attendance_actual_next_week = [att for att in attendances if att.week_type is False or int(att.week_type) != start_week_type]
            attendance_filtred = [att for att in attendance_actual_week if int(att.dayofweek) >= request_date_from.weekday()]
            attendance_filtred += list(attendance_actual_next_week)
            attendance_filtred += list(attendance_actual_week)
            end_week_type = self.env['resource.calendar.attendance'].get_week_type(request_date_to)
            attendance_actual_week = [att for att in attendances if att.week_type is False or int(att.week_type) == end_week_type]
            attendance_actual_next_week = [att for att in attendances if att.week_type is False or int(att.week_type) != end_week_type]
            attendance_filtred_reversed = list(reversed([att for att in attendance_actual_week if int(att.dayofweek) <= request_date_to.weekday()]))
            attendance_filtred_reversed += list(reversed(attendance_actual_next_week))
            attendance_filtred_reversed += list(reversed(attendance_actual_week))

            # find first attendance coming after first_day
            attendance_from = attendance_filtred[0]
            # find last attendance coming before last_day
            attendance_to = attendance_filtred_reversed[0]
        else:
            # find first attendance coming after first_day
            attendance_from = next((att for att in attendances if int(att.dayofweek) >= request_date_from.weekday()), attendances[0] if attendances else default_value)
            # find last attendance coming before last_day
            attendance_to = next((att for att in reversed(attendances) if int(att.dayofweek) <= request_date_to.weekday()), attendances[-1] if attendances else default_value)

        return (attendance_from, attendance_to)
